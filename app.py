# ==============================================================================
# File: app.py
# L'interfaccia utente principale per l'applicazione di generazione di giudizi.
# Utilizza Streamlit per creare un'interfaccia web interattiva.
# Questo file gestisce l'intero flusso di lavoro:
# 1. Caricamento dei file di addestramento e/o del modello fine-tunato.
# 2. Addestramento (fine-tuning) del modello.
# 3. Caricamento del file Excel da completare.
# 4. Generazione dei giudizi per il file caricato.
# 5. Download del file Excel completato.
#
# Logica dei Troncamenti e Pulizia:
# - Troncamento righe: basato sulla colonna 'pos'. La lettura si interrompe
#   quando la sequenza numerica si rompe.
# - Troncamento colonne: vengono rimosse le colonne vuote consecutive da destra.
# - Pulizia finale: vengono rimosse le righe con giudizi vuoti.
#
# Gestione dei Fogli:
# - Vengono letti tutti i fogli tranne 'Prototipo' e 'Medie'.
#
# Identificazione Colonne:
# - "Giudizio" e "Descrizione" sono cercate in modo flessibile.
# - Esclusione di "pos", "Alunno", "assenti", "CNT" dal prompt di input.
# ==============================================================================

# SEZIONE 0: LIBRERIE NECESSARIE E CONFIGURAZIONE
# ==============================================================================
import streamlit as st
import pandas as pd
import os
import shutil
import warnings
import traceback
from datetime import datetime
from io import BytesIO
import json
import time
import requests
import re
import random 
import openpyxl
from collections import Counter
from openpyxl import Workbook, load_workbook

# ==============================================================================
# SEZIONE A: FUNZIONI PER LA GESTIONE DEI DATI EXCEL (basate su excel_reader_v2.py)
# ==============================================================================
class ExcelReader:
    """
    Classe per leggere e preparare dati da file Excel, implementando la logica di
    troncamento, pulizia e identificazione flessibile delle colonne.
    """
    def find_header_row_and_columns(self, df):
        """
        Trova la riga di intestazione e le posizioni delle colonne rilevanti.
        Cerca in modo flessibile 'descrizione' e 'giudizio' nelle prime 10 righe.
        """
        try:
            for i in range(min(10, len(df))):
                row = df.iloc[i].astype(str).str.lower().tolist()
                row_str = " ".join(row)
                
                if re.search(r'\bdescrizione\b', row_str) and re.search(r'\bgiudizio\b', row_str):
                    header_row_index = i
                    headers = df.iloc[header_row_index].astype(str).tolist()
                    
                    description_col_name = next((h for h in headers if re.search(r'\bdescrizione\b', str(h).lower())), None)
                    judgment_col_name = next((h for h in headers if re.search(r'\bgiudizio\b', str(h).lower())), None)
                    pos_col_name = next((h for h in headers if re.search(r'\bpos\b', str(h).lower())), None)

                    if description_col_name and judgment_col_name:
                        return header_row_index, description_col_name, judgment_col_name, pos_col_name
            
            # Fallback a colonna H (indice 7) per il giudizio se la ricerca testuale fallisce
            header_row_index = 0
            headers = df.iloc[header_row_index].astype(str).tolist()
            if len(headers) > 7:
                 judgment_col_name = headers[7]
                 return header_row_index, None, judgment_col_name, None # Non abbiamo trovato la descrizione, ma abbiamo un fallback
            
            return None, None, None, None
        except Exception as e:
            raise Exception(f"Errore nella ricerca dell'intestazione: {e}")
            
    def trim_dataframe_by_empty_columns(self, df):
        """
        Rimuove le colonne vuote da destra finché non ne trova due consecutive piene.
        """
        if df.empty:
            return df
        
        last_col = len(df.columns) - 1
        empty_count = 0
        
        for i in range(last_col, -1, -1):
            if df.iloc[:, i].isnull().all():
                empty_count += 1
            else:
                empty_count = 0
            
            if empty_count >= 2:
                return df.iloc[:, :i+2]
        
        return df
        
    def read_and_prepare_data_from_excel(self, file_object, sheet_names, progress_container):
        """
        Legge i dati dai fogli specificati, prepara un dataframe con 'input_text' e 'target_text'.
        Applica la logica di troncamento e pulizia.
        """
        corpus_list = []
        excluded_sheets = ["Prototipo", "Medie"]
        
        for sheet in sheet_names:
            if sheet in excluded_sheets:
                progress_container(f"FOGLIO ESCLUSO: '{sheet}'.", "warning")
                continue

            progress_container(f"Lettura del foglio: '{sheet}'...", "info")
            file_object.seek(0)
            
            try:
                df = pd.read_excel(file_object, sheet_name=sheet, header=None)
            except Exception as e:
                progress_container(f"Impossibile leggere il foglio '{sheet}'. Ignorato. Errore: {e}", "warning")
                continue

            try:
                header_row_index, description_col, judgment_col, pos_col = self.find_header_row_and_columns(df)
            except Exception as e:
                progress_container(f"Errore nella ricerca dell'intestazione del foglio '{sheet}': {e}", "error")
                continue
            
            if header_row_index is None:
                progress_container(f"Attenzione: Intestazione non trovata nel foglio '{sheet}'. Saltato.", "warning")
                continue
            
            df.columns = df.iloc[header_row_index]
            df = df.iloc[header_row_index + 1:].reset_index(drop=True)

            # Tronca le righe basandosi sulla colonna 'pos'
            if pos_col:
                try:
                    pos_df = df[pos_col].dropna().astype(str).str.strip().str.lower()
                    last_valid_pos_idx = -1
                    for idx, pos_val in enumerate(pos_df):
                        if pos_val.isdigit():
                            last_valid_pos_idx = idx
                        else:
                            break
                    df = df.iloc[:last_valid_pos_idx + 1]
                    progress_container(f"Troncamento righe completato per il foglio '{sheet}'.", "info")
                except Exception as e:
                    progress_container(f"Errore durante il troncamento righe per la colonna 'pos': {e}", "warning")
            
            # Tronca le colonne vuote
            df = self.trim_dataframe_by_empty_columns(df)
            progress_container(f"Troncamento colonne completato per il foglio '{sheet}'.", "info")

            if judgment_col not in df.columns:
                progress_container(f"Attenzione: Colonna 'Giudizio' non trovata nel foglio '{sheet}'. Saltato.", "warning")
                continue

            data_for_dataset = []
            
            # Esclusione delle colonne non pertinenti
            excluded_cols = ['pos', 'alunno', 'assenti', 'cnt']
            
            for _, row in df.iterrows():
                row_dict = row.to_dict()
                target_text = str(row_dict.get(judgment_col, '')).strip()

                if pd.notna(target_text) and target_text:
                    input_parts = []
                    for key, value in row_dict.items():
                        if pd.notna(value) and str(key).strip().lower() not in excluded_cols and str(key).strip().lower() != str(judgment_col).strip().lower():
                            input_parts.append(f"{key}: {value}")
                    
                    input_text = " ".join(input_parts)
                    
                    if input_text:
                        data_for_dataset.append({
                            'input_text': input_text,
                            'target_text': target_text
                        })

            if not data_for_dataset:
                progress_container(f"Attenzione: Nessun dato valido trovato nel foglio '{sheet}'. Saltato.", "warning")
                continue
            
            corpus_list.extend(data_for_dataset)
        
        if not corpus_list:
            progress_container("Nessun dato valido trovato in tutti i fogli del file.", "error")
            return pd.DataFrame()
        
        return pd.DataFrame(corpus_list)

    def get_excel_sheet_names(self, file_object):
        """
        Restituisce una lista con i nomi di tutti i fogli di lavoro in un file Excel.
        """
        try:
            file_object.seek(0)
            workbook = openpyxl.load_workbook(file_object, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            progress_container(f"Errore nella lettura dei nomi dei fogli: {e}", "error")
            return []

# ==============================================================================
# SEZIONE B: CLASSI MOCK PER ADDESTRAMENTO E GENERAZIONE
# ==============================================================================

class MockModel:
    def __init__(self):
        pass

class MockTokenizer:
    def __init__(self):
        pass

class MockModelTrainer:
    def train_model(self, corpus_df, progress_container):
        progress_container("MOCK: Inizio addestramento del modello...", "info")
        progress_container(f"MOCK: Dimensione corpus: {len(corpus_df)} righe.", "info")
        
        for i in range(1, 6):
            time.sleep(0.5)
            progress_container(f"MOCK: Addestramento - Fase {i}/5...", "info")
            progress_container(f"MOCK: Loss: {random.uniform(0.1, 0.5):.4f}", "info")
        
        progress_container("MOCK: Addestramento completato. Salvataggio del checkpoint...", "success")
        time.sleep(1)
        progress_container("MOCK: Checkpoint salvato.", "success")
        return MockModel(), MockTokenizer()

    def load_fine_tuned_model(self, progress_container):
        progress_container("MOCK: Caricamento del modello esistente...", "info")
        time.sleep(1)
        # Simula il caso in cui il modello non è sempre disponibile
        if random.choice([True, True, True, False]):
             progress_container("MOCK: Modello esistente caricato.", "success")
             return MockModel(), MockTokenizer()
        else:
            progress_container("MOCK: Nessun modello esistente trovato.", "error")
            return None, None

    def delete_model(self, progress_container):
        progress_container("MOCK: Eliminazione del modello...", "info")
        time.sleep(1)
        progress_container("MOCK: Modello eliminato.", "success")

class MockJudgmentGenerator:
    def generate_judgments(self, df, model, tokenizer, sheet_name, progress_container):
        progress_container("MOCK: Generazione dei giudizi in corso...", "info")
        
        df_processed = df.copy()
        generated_judgments = []
        
        for _, row in df_processed.iterrows():
            input_text = row['input_text']
            
            # Simulazione della logica di chunking per testi lunghi
            if len(input_text) > 100:
                progress_container(f"MOCK: Testo di input troppo lungo. Applico il 'chunking'...", "info")
                chunk = input_text[:100] + "..."
                generated_judgment = f"Giudizio generato (chunked): {chunk}"
            else:
                generated_judgment = f"Giudizio generato per: {input_text}"
            
            generated_judgments.append(generated_judgment)
        
        # Aggiungo la nuova colonna al DataFrame
        df_processed['Giudizio_Generato'] = generated_judgments
        
        progress_container("MOCK: Generazione completata!", "success")
        return df_processed

class CorpusBuilder:
    def build_or_update_corpus(self, new_df, progress_container):
        progress_container("Aggiornamento del corpus...", "info")
        if st.session_state.corpus_df.empty:
            st.session_state.corpus_df = new_df
        else:
            # Unisce il nuovo DataFrame con quello esistente e rimuove i duplicati
            st.session_state.corpus_df = pd.concat([st.session_state.corpus_df, new_df], ignore_index=True)
            st.session_state.corpus_df.drop_duplicates(subset=['input_text'], keep='first', inplace=True)
        
        progress_container(f"Corpus aggiornato. Righe totali: {len(st.session_state.corpus_df)}", "success")
        return st.session_state.corpus_df

    def delete_corpus(self, progress_container):
        progress_container("MOCK: Eliminazione del corpus...", "info")
        time.sleep(1)
        if 'corpus_df' in st.session_state:
            st.session_state.corpus_df = pd.DataFrame()
        progress_container("MOCK: Corpus eliminato.", "success")
        
# Istanzio le classi
er = ExcelReader()
mt = MockModelTrainer()
jg = MockJudgmentGenerator()
cb = CorpusBuilder()

# Ignoriamo i FutureWarning per mantenere la console pulita.
warnings.filterwarnings("ignore")

# Definizione della variabile dell'API Key per l'uso con le API di Gemini
GEMINI_API_KEY = ""

# ==============================================================================
# SEZIONE 1: FUNZIONI AUSILIARIE
# ==============================================================================

def progress_container(message, type):
    """
    Gestisce l'aggiornamento del placeholder di stato con un messaggio.
    """
    if "status_placeholder" not in st.session_state:
        st.session_state.status_placeholder = st.empty()
    
    if type == "info":
        st.session_state.status_placeholder.info(message)
    elif type == "success":
        st.session_state.status_placeholder.success(message)
    elif type == "warning":
        st.session_state.status_placeholder.warning(message)
    elif type == "error":
        st.session_state.status_placeholder.error(message)

def reset_project_state():
    """
    Resetta tutti i file del progetto (corpus, modello) e lo stato di sessione.
    """
    try:
        # Elimina il corpus (mock)
        cb.delete_corpus(lambda msg, type: progress_container(msg, type))
        # Elimina il modello (mock)
        mt.delete_model(lambda msg, type: progress_container(msg, type))
        progress_container("Progetto resettato. Tutti i dati sono stati eliminati.", "success")
        
        # Reset dello stato di sessione
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        
    except Exception as e:
        progress_container(f"Errore durante il reset del progetto: {e}", "error")
        st.error(f"Errore: {traceback.format_exc()}")
        
    st.experimental_rerun()
    
def get_session_id():
    """
    Genera un ID di sessione per distinguere i file in un ambiente multi-utente.
    """
    if "session_id" not in st.session_state:
        st.session_state.session_id = datetime.now().strftime("%Y%m%d%H%M%S")
    return st.session_state.session_id

# ==============================================================================
# SEZIONE 2: INTERFACCIA UTENTE E LOGICA PRINCIPALE
# ==============================================================================

# Impostazioni della pagina
st.set_page_config(
    page_title="Giudizi-AI",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inizializzazione di Streamlit session state per mantenere lo stato dell'app
# Inizializzazione Condizionale per evitare errori all'avvio
if 'corpus_df' not in st.session_state:
    st.session_state.corpus_df = pd.DataFrame()
if 'model_ready' not in st.session_state:
    st.session_state.model_ready = False
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = None
if 'selected_sheet_trainer' not in st.session_state:
    st.session_state.selected_sheet_trainer = None
if 'uploaded_process_file' not in st.session_state:
    st.session_state.uploaded_process_file = None
if 'selected_sheet_process' not in st.session_state:
    st.session_state.selected_sheet_process = None
if 'process_completed_file' not in st.session_state:
    st.session_state.process_completed_file = None
if 'status_placeholder' not in st.session_state:
    st.session_state.status_placeholder = st.empty()


# Titolo e descrizione dell'app
st.title("🤖 Giudizi-AI: Generatore di Giudizi 🤖")
st.markdown("---")
st.markdown("Questa applicazione ti aiuta a generare automaticamente giudizi per le tue valutazioni. Segui i passaggi di seguito.")

# Layout in colonne per la sidebar e il contenuto principale
main_col, sidebar_col = st.columns([3, 1])

with sidebar_col:
    st.markdown("### Impostazioni e Stato")
    
    st.info("Utilizza questa sidebar per resettare l'applicazione o per informazioni di stato.")
    
    # Bottone per caricare il modello
    if st.button("Carica Modello Esistente"):
        try:
            model, tokenizer = mt.load_fine_tuned_model(lambda msg, type: progress_container(msg, type))
            if model and tokenizer:
                st.session_state.model_ready = True
                st.session_state.model = model
                st.session_state.tokenizer = tokenizer
                progress_container("Modello caricato con successo.", "success")
            else:
                st.session_state.model_ready = False
                progress_container("Nessun modello trovato. Devi addestrarne uno.", "error")
        except Exception as e:
            progress_container(f"Errore durante il caricamento del modello: {e}", "error")
            st.error(f"Errore: {traceback.format_exc()}")
            
    # Bottone per resettare lo stato del progetto
    if st.button("Resetta tutto il progetto"):
        reset_project_state()
        
    st.markdown("---")
    st.write("### Stato Attuale")
    if st.session_state.model_ready:
        st.success("✅ Modello caricato e pronto!")
    else:
        st.error("❌ Nessun modello caricato.")
        
    st.write(f"Righe nel corpus: {len(st.session_state.corpus_df)}")
    
    # Mostra l'ID della sessione
    st.text(f"ID Sessione: {get_session_id()}")

# ==============================================================================
# SEZIONE 1: ADDESTRAMENTO DEL MODELLO
# ==============================================================================
with main_col:
    st.header("1. Addestramento del Modello 🧠")
    st.info("Carica uno o più file Excel per creare o aggiornare il corpus di addestramento. Le colonne 'Descrizione' e 'Giudizio' verranno utilizzate per il fine-tuning del modello.")
    
    # Placeholder per i messaggi di stato
    st.session_state.status_placeholder = st.empty()
    
    # Caricamento dei file per l'addestramento
    uploaded_files_trainer = st.file_uploader(
        "Carica i file di addestramento (.xlsx, .xls, .xlsm)",
        type=["xlsx", "xls", "xlsm"],
        accept_multiple_files=True,
        key="uploader_trainer"
    )
    
    # Aggiorno lo stato dei file caricati
    if uploaded_files_trainer:
        st.session_state.uploaded_files = uploaded_files_trainer
        try:
            first_file = uploaded_files_trainer[0]
            sheet_names = er.get_excel_sheet_names(first_file)
            st.session_state.selected_sheet_trainer = st.selectbox(
                "Seleziona il foglio di lavoro da utilizzare per l'addestramento:",
                sheet_names,
                key="sheet_select_trainer"
            )
        except Exception as e:
            progress_container(f"Errore nella lettura dei fogli del file: {e}", "error")
            st.error(f"Errore: {traceback.format_exc()}")
        
    if st.button("Avvia Addestramento"):
        if st.session_state.uploaded_files:
            try:
                progress_container("Lettura dei file e preparazione del corpus...", "info")
                new_df = pd.DataFrame()
                for file in st.session_state.uploaded_files:
                    if st.session_state.selected_sheet_trainer:
                        df_temp = er.read_and_prepare_data_from_excel(file, [st.session_state.selected_sheet_trainer], lambda msg, type: progress_container(msg, type))
                        new_df = pd.concat([new_df, df_temp], ignore_index=True)
                    else:
                        progress_container("Seleziona un foglio di lavoro.", "warning")
                        continue

                st.session_state.corpus_df = cb.build_or_update_corpus(new_df, lambda msg, type: progress_container(msg, type))
                
                if not st.session_state.corpus_df.empty:
                    st.session_state.model, st.session_state.tokenizer = mt.train_model(st.session_state.corpus_df, lambda msg, type: progress_container(msg, type))
                    if st.session_state.model and st.session_state.tokenizer:
                        st.session_state.model_ready = True
                        progress_container("Addestramento completato con successo!", "success")
                        st.balloons()
                    else:
                        st.session_state.model_ready = False
                        progress_container("Errore durante l'addestramento. Controlla i dati.", "error")
                else:
                    progress_container("Corpus vuoto. Impossibile avviare l'addestramento.", "error")
            except Exception as e:
                progress_container(f"Errore durante il processo di addestramento: {e}", "error")
                st.error(f"Errore: {traceback.format_exc()}")
        else:
            st.warning("Per addestrare il modello, devi prima caricare almeno un file e selezionare un foglio.")

st.markdown("---")

# ==============================================================================
# SEZIONE 2: GENERAZIONE DEI GIUDIZI
# ==============================================================================
with main_col:
    st.header("2. Generazione dei Giudizi 🤖")
    st.info("Carica il file Excel che vuoi completare. L'applicazione genererà i giudizi per le righe mancanti.")
    
    # Caricamento del file per la generazione
    uploaded_process_file = st.file_uploader(
        "Carica il file da completare (.xlsx, .xls, .xlsm)",
        type=["xlsx", "xls", "xlsm"],
        key="uploader_process"
    )
    
    if uploaded_process_file:
        st.session_state.uploaded_process_file = uploaded_process_file
        try:
            sheet_names = er.get_excel_sheet_names(uploaded_process_file)
            st.session_state.selected_sheet_process = st.selectbox(
                "Seleziona il foglio di lavoro da completare:",
                sheet_names,
                key="sheet_select_process"
            )
        except Exception as e:
            progress_container(f"Errore nella lettura dei fogli del file: {e}", "error")
            st.error(f"Errore: {traceback.format_exc()}")
        
    if st.button("Genera Giudizi"):
        if st.session_state.model_ready:
            if st.session_state.uploaded_process_file and st.session_state.selected_sheet_process:
                try:
                    progress_container("Lettura del file...", "info")
                    df = er.read_and_prepare_data_from_excel(st.session_state.uploaded_process_file, [st.session_state.selected_sheet_process], lambda msg, type: progress_container(msg, type))
                    
                    if not df.empty:
                        progress_container("Generazione dei giudizi in corso...", "info")
                        df_with_judgments = jg.generate_judgments(df, st.session_state.model, st.session_state.tokenizer, st.session_state.selected_sheet_process, lambda msg, type: progress_container(msg, type))
                        
                        st.session_state.process_completed_file = df_with_judgments
                        st.session_state.selected_sheet = st.session_state.selected_sheet_process
                        progress_container("Generazione completata!", "success")
                    else:
                        progress_container("Il file caricato è vuoto o non contiene dati validi.", "error")
                except Exception as e:
                    progress_container(f"Errore nella generazione dei giudizi: {e}", "error")
                    st.error(f"Errore: {traceback.format_exc()}")
            else:
                st.warning("Per generare i giudizi, devi prima caricare un file nella sezione '2. Generazione dei Giudizi'.")
        else:
            st.warning("Per generare i giudizi, devi prima addestrare un modello nella sezione '1. Addestramento del Modello'.")

st.markdown("---")

# ==============================================================================
# SEZIONE 3: VISUALIZZAZIONE RISULTATI E DOWNLOAD
# ==============================================================================
with main_col:
    st.header("3. Stato e Download")

    if st.session_state.process_completed_file is not None:
        st.write("### Scarica il file completato")
        st.write("Di seguito, puoi vedere l'anteprima del file con i giudizi generati e scaricare il documento completo.")
        
        st.dataframe(st.session_state.process_completed_file)
        
        output_buffer = BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            st.session_state.process_completed_file.to_excel(writer, index=False, sheet_name=st.session_state.selected_sheet)
        output_buffer.seek(0)
        
        st.download_button(
            label="Scarica il file aggiornato",
            data=output_buffer,
            file_name=f"Giudizi_Generati_{st.session_state.selected_sheet}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("I risultati appariranno qui una volta che avrai generato i giudizi.")
