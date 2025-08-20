# ==============================================================================
# File: excel_reader.py
# Logica per la preparazione dei dati da file Excel, integrando le
# funzionalità di '33 Funziona.txt' per una lettura più robusta.
# ==============================================================================
import pandas as pd
import openpyxl
import re
from io import BytesIO
import traceback
import os
import shutil
import json
from datetime import datetime

# ==============================================================================
# SEZIONE 1: FUNZIONI AUSILIARIE
# ==============================================================================

def make_columns_unique(columns):
    """
    Garantisce che i nomi delle colonne siano unici, aggiungendo un contatore
    se necessario.
    """
    seen = {}
    new_columns = []
    for col in columns:
        original_col = col
        if original_col in seen:
            seen[original_col] += 1
            new_columns.append(f"{original_col}_{seen[original_col]}")
        else:
            seen[original_col] = 0
            new_columns.append(original_col)
    return new_columns

def find_header_row_and_columns(df):
    """
    Trova la riga di intestazione e le posizioni delle colonne 'Giudizio' e 'pos'.
    Scansiona le prime 50 righe per trovare l'intestazione.
    """
    giudizio_col_name = None
    pos_col_name = None
    header_row_index = None

    # Normalizziamo i nomi delle colonne da cercare per una ricerca case-insensitive
    target_giudizio = re.compile(r"giudizio|valutazione|descrizione", re.IGNORECASE)
    target_pos = re.compile(r"pos|posizione|numero", re.IGNORECASE)

    # Scansiona le prime 50 righe per trovare l'intestazione
    for i in range(min(50, len(df))):
        row = df.iloc[i].astype(str)
        # Cerchiamo la colonna 'Giudizio' nell'intestazione
        if any(re.search(target_giudizio, str(cell)) for cell in row):
            header_row_index = i
            break
    
    if header_row_index is None:
        raise ValueError("Impossibile trovare la riga di intestazione che contiene la parola 'Giudizio' o un suo sinonimo.")

    # Usiamo la riga trovata come intestazione temporanea per cercare le colonne
    header_row = df.iloc[header_row_index].astype(str)
    
    # Cerchiamo la colonna 'Giudizio'
    matches_giudizio = [re.search(target_giudizio, col) for col in header_row]
    if any(matches_giudizio):
        giudizio_col_name = header_row[matches_giudizio.index(next(m for m in matches_giudizio if m))]
    
    # Cerchiamo la colonna 'pos'
    matches_pos = [re.search(target_pos, col) for col in header_row]
    if any(matches_pos):
        pos_col_name = header_row[matches_pos.index(next(m for m in matches_pos if m))]
        
    # Fallback per la colonna 'Giudizio'
    if not giudizio_col_name:
        # Assumiamo che la colonna H (indice 7) sia la destinazione
        if 7 < len(df.columns):
            giudizio_col_name = df.columns[7]
        else:
            raise ValueError("Impossibile trovare la colonna 'Giudizio' e la colonna H non esiste.")
    
    if not pos_col_name:
        # Prova a trovare la colonna 'pos' in base al suo contenuto
        for col in df.columns:
            # Controllo se le prime 50 righe di questa colonna contengono numeri sequenziali
            try:
                values = pd.to_numeric(df[col].iloc[:50].dropna(), errors='coerce')
                # Verifichiamo che ci siano più di un valore unico e che siano in ordine
                if values.nunique() > 1 and all(values == values.sort_values()):
                    pos_col_name = col
                    break
            except (ValueError, TypeError):
                continue
    
    return header_row_index, giudizio_col_name, pos_col_name

def trim_dataframe_by_numeric_id_column(df, pos_col_name):
    """
    Tronca il DataFrame in base alla sequenza numerica nella colonna 'pos'.
    """
    if pos_col_name is None or pos_col_name not in df.columns:
        return df

    # Converti la colonna 'pos' in tipo numerico, gestendo gli errori
    df['temp_pos'] = pd.to_numeric(df[pos_col_name], errors='coerce')
    
    # Rimuovi le righe dove la colonna 'pos' non è un numero valido
    df = df.dropna(subset=['temp_pos'])
    
    if df.empty:
        return df

    # Trova l'ultimo indice dove la sequenza numerica è consecutiva
    max_idx = 0
    start_num = int(df['temp_pos'].iloc[0])
    for i in range(len(df)):
        if int(df['temp_pos'].iloc[i]) == start_num + i:
            max_idx = i
        else:
            break
            
    df = df.iloc[:max_idx + 1]
    df = df.drop(columns=['temp_pos'])
    return df

def trim_dataframe_by_empty_columns(df):
    """
    Rimuove le colonne vuote consecutive dalla destra del DataFrame.
    """
    cols_to_drop = []
    empty_count = 0
    # Itera sulle colonne da destra a sinistra
    for col in reversed(df.columns):
        if df[col].isnull().all():
            empty_count += 1
            cols_to_drop.append(col)
        else:
            empty_count = 0
        if empty_count >= 2:
            break
            
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    return df

def read_and_prepare_data_from_excel(file_object, progress_container, sheets_to_ignore=['prototipo', 'medie']):
    """
    Legge tutti i fogli di lavoro di un file Excel e prepara i dati per l'addestramento.
    """
    corpus_list = []
    
    try:
        file_object.seek(0)
        workbook = openpyxl.load_workbook(file_object, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        # Filtra i fogli da ignorare
        sheets_to_process = [s for s in sheet_names if s.lower().strip() not in [sh.lower().strip() for sh in sheets_to_ignore]]
        
        if not sheets_to_process:
            progress_container("Nessun foglio di lavoro valido trovato da processare.", "warning")
            return pd.DataFrame()
            
        progress_container(f"Trovati {len(sheets_to_process)} fogli da processare: {', '.join(sheets_to_process)}", "info")
        
        # Legge il file di nuovo con pandas per processare ogni foglio
        file_object.seek(0)
        all_sheets = pd.read_excel(file_object, sheet_name=None)
        
        for sheet in sheets_to_process:
            try:
                df = all_sheets[sheet]
                if df.empty:
                    progress_container(f"Attenzione: Il foglio '{sheet}' è vuoto. Saltato.", "warning")
                    continue
                
                # Rimuove le colonne completamente vuote
                df.dropna(axis=1, how='all', inplace=True)
                if df.empty:
                    progress_container(f"Attenzione: Il foglio '{sheet}' è vuoto dopo la pulizia. Saltato.", "warning")
                    continue

                progress_container(f"Elaborazione del foglio '{sheet}'...", "info")
                
                # Trova la riga di intestazione e le colonne chiave
                header_row_index, giudizio_col_name, pos_col_name = find_header_row_and_columns(df.copy())
                
                # Imposta la riga di intestazione
                df.columns = make_columns_unique(df.iloc[header_row_index].astype(str).str.strip().tolist())
                df = df.iloc[header_row_index + 1:].reset_index(drop=True)

                # Tronca il DataFrame in base alla colonna 'pos'
                if pos_col_name and pos_col_name in df.columns:
                    df = trim_dataframe_by_numeric_id_column(df, pos_col_name)

                # Tronca le colonne vuote dalla destra
                df = trim_dataframe_by_empty_columns(df)
                
                # Rimuovi righe diventate vuote dopo il troncamento
                df.dropna(how='all', inplace=True)
                
                if df.empty:
                    progress_container(f"Attenzione: Nessun dato valido trovato nel foglio '{sheet}' dopo la pulizia. Saltato.", "warning")
                    continue
                
                # Preparazione dei dati per il fine-tuning
                data_for_dataset = []
                # Crea una lista di colonne da escludere, in modo case-insensitive
                cols_to_exclude = {col.lower().strip() for col in ['alunno', 'assenti', 'cnt', 'pos', 'giudizio', pos_col_name]}

                # Crea i prompt e i target
                for _, row in df.iterrows():
                    input_text_parts = []
                    # Unisci i dati di tutte le colonne utili nel prompt
                    for col_name, value in row.items():
                        if col_name.lower().strip() not in cols_to_exclude and pd.notna(value):
                            input_text_parts.append(f"{col_name}: {value}")
                    
                    input_text = " ".join(input_text_parts)
                    target_text = str(row[giudizio_col_name])
                    
                    if input_text and target_text.strip():
                        data_for_dataset.append({
                            'input_text': input_text,
                            'target_text': target_text
                        })

                if not data_for_dataset:
                    progress_container(f"Attenzione: Nessun dato valido trovato nel foglio '{sheet}'. Saltato.", "warning")
                    continue
                
                corpus_list.extend(data_for_dataset)
            
            except Exception as e:
                progress_container(f"Errore nella lettura del foglio '{sheet}': {e}", "error")
                progress_container(f"Traceback: {traceback.format_exc()}", "error")
        
        if not corpus_list:
            progress_container("Nessun dato valido trovato in tutti i fogli del file.", "error")
            return pd.DataFrame()
            
        return pd.DataFrame(corpus_list)

    except Exception as e:
        progress_container(f"Errore nella lettura del file: {e}", "error")
        progress_container(f"Traceback: {traceback.format_exc()}", "error")
        return pd.DataFrame()


def read_single_sheet(file_object, sheet_name, progress_container):
    """
    Legge e prepara i dati da un singolo foglio di lavoro.
    """
    try:
        file_object.seek(0)
        df = pd.read_excel(file_object, sheet_name=sheet_name)
        
        if df.empty:
            progress_container(f"Attenzione: Il foglio '{sheet_name}' è vuoto. Impossibile generare giudizi.", "warning")
            return None
        
        df.dropna(axis=1, how='all', inplace=True)
        if df.empty:
            progress_container(f"Attenzione: Il foglio '{sheet_name}' è vuoto dopo la pulizia. Impossibile generare giudizi.", "warning")
            return None

        # Trova la riga di intestazione e le colonne chiave
        header_row_index, giudizio_col_name, pos_col_name = find_header_row_and_columns(df.copy())
        
        # Imposta la riga di intestazione
        df.columns = make_columns_unique(df.iloc[header_row_index].astype(str).str.strip().tolist())
        df = df.iloc[header_row_index + 1:].reset_index(drop=True)
        
        # Tronca il DataFrame in base alla colonna 'pos'
        if pos_col_name and pos_col_name in df.columns:
            df = trim_dataframe_by_numeric_id_column(df, pos_col_name)

        # Tronca le colonne vuote dalla destra
        df = trim_dataframe_by_empty_columns(df)
        
        # Rimuovi righe diventate vuote dopo il troncamento
        df.dropna(how='all', inplace=True)

        if df.empty:
            progress_container(f"Nessun dato valido trovato nel foglio '{sheet_name}'.", "error")
            return None

        return df

    except Exception as e:
        progress_container(f"Errore nella lettura del file: {e}", "error")
        progress_container(f"Traceback: {traceback.format_exc()}", "error")
        return None

def get_excel_sheet_names(file_object):
    """
    Restituisce una lista con i nomi di tutti i fogli di lavoro in un file Excel.
    """
    try:
        file_object.seek(0) # Riporta il puntatore all'inizio del file
        workbook = openpyxl.load_workbook(file_object, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names
    except Exception as e:
        st.error(f"Errore nella lettura dei nomi dei fogli: {e}")
        return []

